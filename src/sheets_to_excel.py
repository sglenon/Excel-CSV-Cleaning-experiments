"""
Excel Sheet Splitter

Splits each sheet of an Excel (.xlsx) file into a separate Excel file using openpyxl.
Preserves cell formatting, formulas, merged cells, and styles, but may not preserve charts or images.
Works on Windows, macOS, and Linux.

Usage:
    python v3_sheets_to_excel.py <input_excel_file.xlsx> <output_folder>

Args:
    input_excel_file.xlsx: Path to the source Excel file.
    output_folder: Directory where the separated sheet files will be saved.

Dependencies:
    - openpyxl

Example:
    python v3_sheets_to_excel.py "MyWorkbook.xlsx" "sheets_to_excel/"

"""

import openpyxl
import os
import sys
import argparse

def separate_sheets_with_openpyxl(input_file, output_folder):
    """
    Split each sheet of an Excel file into a new workbook using openpyxl.

    Each output file contains a single sheet from the original workbook, preserving formatting,
    formulas, merged cells, and styles. Charts and images may not be preserved.

    Args:
        input_file (str): Path to the source Excel file (.xlsx).
        output_folder (str): Directory where the separated sheet files will be saved.

    Raises:
        SystemExit: If the input file does not exist or output directory cannot be created.
    """
    if not os.path.exists(input_file):
        print(f"❌ Error: The input file was not found at '{input_file}'")
        sys.exit(1)

    try:
        os.makedirs(output_folder, exist_ok=True)
        print(f"✅ Output folder '{output_folder}' is ready.")
    except OSError as e:
        print(f"❌ Error: Could not create output directory '{output_folder}': {e}")
        sys.exit(1)

    try:
        source_wb = openpyxl.load_workbook(input_file)
    except Exception as e:
        print(f"❌ Error: Failed to read the Excel file '{input_file}': {e}")
        sys.exit(1)

    base_name = os.path.splitext(os.path.basename(input_file))[0]
    print(f"\nProcessing '{os.path.basename(input_file)}'...")

    for idx, sheet_name in enumerate(source_wb.sheetnames, start=1):
        print(f"  - Processing sheet {idx}: '{sheet_name}'")

        # Create a new workbook for the sheet
        new_wb = openpyxl.Workbook()
        default_sheet = new_wb.active
        new_wb.remove(default_sheet)

        # Get the source sheet
        source_sheet = source_wb[sheet_name]

        # Create a new sheet in the new workbook with the same title
        new_sheet = new_wb.create_sheet(title=sheet_name)

        # Copy data and formatting from the source to the new sheet
        for row in source_sheet.iter_rows():
            for cell in row:
                new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    # Copy cell font
                    new_cell.font = openpyxl.styles.Font(
                        name=cell.font.name,
                        size=cell.font.size,
                        bold=cell.font.bold,
                        italic=cell.font.italic,
                        color=cell.font.color,
                    )
                    # Copy cell fill
                    new_cell.fill = openpyxl.styles.PatternFill(
                        fill_type=cell.fill.fill_type,
                        start_color=cell.fill.start_color,
                        end_color=cell.fill.end_color,
                    )
                    # Copy cell border
                    new_cell.border = openpyxl.styles.Border(
                        left=cell.border.left,
                        right=cell.border.right,
                        top=cell.border.top,
                        bottom=cell.border.bottom,
                    )
                    # Copy cell alignment
                    new_cell.alignment = openpyxl.styles.Alignment(
                        horizontal=cell.alignment.horizontal,
                        vertical=cell.alignment.vertical,
                        wrap_text=cell.alignment.wrap_text,
                    )
                    # Copy number format
                    new_cell.number_format = cell.number_format

        # Copy merged cells
        for merge_range in source_sheet.merged_cells.ranges:
            new_sheet.merge_cells(str(merge_range))

        # Sanitize sheet name for filename
        safe_sheet_name = "".join([c for c in sheet_name if c.isalnum() or c in (' ', '_', '-')]).rstrip().replace(' ', '_')

        # Create output filename -> {originalfilename}_sheet{idx}_{sheetname}.xlsx
        output_filename = os.path.join(output_folder, f"{base_name}_sheet{idx}_{safe_sheet_name}.xlsx")

        try:
            print(f"    -> Saving to '{output_filename}'")
            new_wb.save(output_filename)
        except Exception as e:
            print(f"    -> ❌ Error saving '{output_filename}': {e}")

    print("\n🎉 Separation complete using openpyxl.")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Split each sheet of an Excel file into a separate Excel file using openpyxl."
    )
    parser.add_argument("input_file", help="Path to the source Excel file (.xlsx).")
    parser.add_argument("output_folder", help="Directory where the separated sheet files will be saved.")
    args = parser.parse_args()
    separate_sheets_with_openpyxl(args.input_file, args.output_folder)
